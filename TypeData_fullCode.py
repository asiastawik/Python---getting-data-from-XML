import xml.etree.ElementTree as ET
import pandas as pd
import numpy as np
import os
import ntpath 
import openpyxl
from openpyxl import load_workbook
from openpyxl.drawing.image import Image as XLImage
from openpyxl.styles import PatternFill, Alignment, GradientFill, Font
from openpyxl.utils import get_column_letter
from openpyxl.styles.borders import Border, Side
import lxml.etree
from PIL import Image as PILImage

def formatting(path):

    #ile kolumn i wierszy chcemy formatować?
    kolumny = 99
    wiersze = 400
    
    #wczytywanie pliku do edycji
    wb = load_workbook(path)
    ws = wb.active
    sheet = wb["Sheet1"] 

    #filtrowanie
    ws.insert_rows(5)
    sheet.auto_filter.ref = "A5:CU300"

    #zmiana rozmiarów wierszy
    sheet.row_dimensions[1].height = 20
    sheet.row_dimensions[2].height = 40
    sheet.row_dimensions[3].height = 40
    sheet.row_dimensions[4].height = 110

    #zmiana rozmiarów kolumn A-E
    for row in range(2, wiersze):
        for col in range(1,6):
            col_letter = get_column_letter(col)
            ws.column_dimensions[col_letter].width = 20

    #formatowanie wyglądu nagłówków A-E
    light_blue = "009999FF"
    for rows in sheet.iter_rows(min_row=4, max_row=5, min_col=1, max_col=5):
        for cell in rows:
            if cell.row:
                cell.fill = PatternFill(start_color=light_blue, end_color=light_blue, fill_type = "solid")
                cell.font = Font(bold=True, size=12)
                cell.alignment = Alignment(horizontal="center", vertical="center")


    #obamowanie
    thin_border = Border(left=Side(style='thin'), 
                         right=Side(style='thin'), 
                         top=Side(style='thin'), 
                         bottom=Side(style='thin'))

    for rows in sheet.iter_rows(min_row=1, max_row=wiersze, min_col=1, max_col=kolumny):
        for cell in rows:
            if cell.row:
                cell.border = thin_border
                cell.alignment = Alignment(horizontal="center", vertical="center")                        

    #nagłówek - obrazy
    ws.merge_cells('A1:E3')
    #zabezpieczenie przed brakiem potrzebnych obrazów w folderze
    try:
        logo = XLImage("./logo.png")
    except FileNotFoundError:
        logo_pil = PILImage.new("RGB", (500, 120), (255, 255, 255))
        logo_pil.save("logo_pil.png")
        logo = XLImage("logo_pil.png")
    else:   
        logo.width = 500
        logo.height = 120
        
    try:
        element = XLImage("./element.png")
    except FileNotFoundError:
        element_pil = PILImage.new("RGB", (140, 140), (255, 255, 255))
        element_pil.save("element_pil.png")
        element = XLImage("element_pil.png")
    else:  
        element.width = 140
        element.height = 140

    element.alignment = Alignment(horizontal="left", vertical="center")
    ws.add_image(logo, 'A1')
    ws.add_image(element, 'E1')

    #formatowanie wyglądu nagłówków wiersza 1
    ws.merge_cells('F1:CO1')
    ws.merge_cells('CP1:CS1')
    ws.merge_cells('CT1:CU1')
    sheet["F1"].value = "Part Number - FA"
    sheet["F1"].alignment = Alignment(horizontal="left", vertical="center")
    sheet["CP1"].value = "- LS"
    sheet["CT1"].value = "Testing"

    dark_blue = "00003366"
    white = "00FFFFFF"
    for rows in sheet.iter_rows(min_row=1, max_row=1, min_col=6, max_col=kolumny):
        for cell in rows:
            if cell.row:
                cell.fill = PatternFill(start_color=dark_blue, end_color=dark_blue, fill_type = "solid")
                cell.font = Font(bold=True, size=15, color=white)

    #formatowanie wyglądu nagłówków wiersza 2
    for row in range(1,wiersze):
        for col in range(6,100):
            col_letter = get_column_letter(col)
            ws.column_dimensions[col_letter].width = 5

    ws.merge_cells('F2:K2')
    sheet["F2"].value = "Machine Tools"
    ws.merge_cells('L2:AD2')
    sheet["L2"].value = "Pos.10 GearHousing"
    ws.merge_cells('AE2:AJ2')
    sheet["AE2"].value = "Pos.014 TieRod"
    sheet["AK2"].value = "Pos.16"
    ws.merge_cells('AL2:AO2')
    sheet["AL2"].value = "Pos.2x Stud"
    ws.merge_cells('AP2:AU2')
    sheet["AP2"].value = "Pos.28 DDU"
    ws.merge_cells('AV2:AX2')
    sheet["AV2"].value = "Pos.034 ValveBody"
    ws.merge_cells('AY2:BD2')
    sheet["AY2"].value = "Pos.36"
    sheet["BE2"].value = "Pos.40"
    ws.merge_cells('BF2:BH2')
    sheet["BF2"].value = "Pos.42 Push Rod assy"
    ws.merge_cells('BI2:BM2')
    sheet["BI2"].value = "Pos.50"
    ws.merge_cells('BN2:BU2')
    sheet["BN2"].value = "Pos.058 PowerPack"
    ws.merge_cells('BV2:BX2')
    sheet["BV2"].value = "Pos.82"
    ws.merge_cells('BY2:CA2')
    sheet["BY2"].value = "Pos.083"
    ws.merge_cells('CB2:CC2')
    sheet["CB2"].value = "Pos.084/085"
    ws.merge_cells('CD2:CG2')
    sheet["CD2"].value = "Pos.088"
    ws.merge_cells('CH2:CL2')
    sheet["CH2"].value = "Pos.094"
    sheet["CM2"].value = "Pos.97"
    ws.merge_cells('CN2:CO2')
    sheet["CN2"].value = "Pos.499"
    ws.merge_cells('CP2:CS2')
    sheet["CP2"].value = "Pos.32"

    blue = "000000FF"
    yellow = "00FFFF00"
    for rows in sheet.iter_rows(min_row=2, max_row=2, min_col=6, max_col=kolumny):
        for cell in rows:
            if cell.row:
                cell.fill = PatternFill(start_color=blue, end_color=blue, fill_type = "solid")
                cell.font = Font(bold=True, size=15, color=yellow)
                cell.alignment = Alignment(wrap_text=True, horizontal='center', vertical='center')

    #formatowanie wyglądu nagłówków wiersza 3,4
    for row in range(1,wiersze):
        for col in range(6,12):
            col_letter = get_column_letter(col)
            ws.column_dimensions[col_letter].width = 5

    light_blue = "009999FF"
    light_yellow = "FFFFE0"
    for rows in sheet.iter_rows(min_row=3, max_row=3, min_col=6, max_col=kolumny):
        for cell in rows:
             if cell.row:
                cell.fill = PatternFill(start_color=light_blue, end_color=light_blue, fill_type = "solid")
                cell.font = Font(bold=True, size=10)
                cell.alignment = Alignment(wrap_text=True, horizontal='center', vertical='center')

    for rows in sheet.iter_rows(min_row=4, max_row=5, min_col=6, max_col=kolumny):
        for cell in rows:
            if cell.row:
                cell.fill = PatternFill(start_color=light_yellow, end_color=light_yellow, fill_type = "solid")
                cell.font = Font(bold=True, size=9)
                cell.alignment = Alignment(wrap_text=True, horizontal='center', text_rotation=90)


    #co drugi wiersz na szaro
    grey = "E8E8E8"
    for rows in sheet.iter_rows(min_row=6, max_row=wiersze, min_col=1, max_col=kolumny):
        for cell in rows:
            if cell.row % 2:
                cell.fill = PatternFill(start_color=grey, end_color=grey, fill_type = "solid")

    #Machine Tools
    ws.merge_cells('F3:G3')
    ws.merge_cells('H3:I3')
    ws.merge_cells('J3:K3')
    sheet["F3"].value = "St. 53 DryTest"
    sheet["H3"].value = "St. 65 CamCheck"
    sheet["J3"].value = "St. 73"

    sheet["F4"].value = "ECU PlugMirror"
    sheet["G4"].value = "ECU PlugAngle"
    sheet["H4"].value = "CamProg_Gasket"
    sheet["I4"].value = "CamProg_Spacer"
    sheet["J4"].value = "Mounting ?"
    sheet.column_dimensions['K'].width = 6
    sheet["K4"].value = "MountingPosition/ToolLifterAngle"

    #Pos.10 GearHousing
    ws.merge_cells('L3:M3')
    #N3
    ws.merge_cells('O3:P3')
    ws.merge_cells('Q3:R3')
    #S3
    ws.merge_cells('T3:V3')
    ws.merge_cells('W3:X3')
    ws.merge_cells('Y3:Z3')
    ws.merge_cells('AA3:AB3')
    ws.merge_cells('AC3:AD3')
    sheet["L3"].value = "Design"
    sheet["N3"].value = "ZAM"
    sheet["O3"].value = "Tierod"
    sheet["Q3"].value = "Riveting Progrs."
    #S3-empty
    sheet["T3"].value = "Typgruppe"
    sheet["W3"].value = "RobPos.21"
    sheet["Y3"].value = "RobPos.22"
    sheet["AA3"].value = "RobPos.23"
    sheet["AC3"].value = "RobPos.24"

    sheet["L4"].value = "Standard"
    sheet["M4"].value = "Mirrored"
    sheet["N4"].value = "ZAM"
    sheet["O4"].value = "100 Bore Distance"
    sheet["P4"].value = "101.8 Bore Distance"
    sheet["Q4"].value = "Program BTR 1"
    sheet["R4"].value = "Program BTR 2"
    sheet["S4"].value = "Number of Stud Bores"
    sheet["T4"].value = "2_Studs_RobPos_21_23"
    sheet["U4"].value = "2_Studs_RobPos_22_24"
    sheet["V4"].value = "4_Studs_RobPos_21-24"
    sheet.column_dimensions['W'].width = 6
    sheet.column_dimensions['X'].width = 6
    sheet.column_dimensions['Y'].width = 6
    sheet.column_dimensions['Z'].width = 6
    sheet.column_dimensions['AA'].width = 6
    sheet.column_dimensions['AB'].width = 6
    sheet.column_dimensions['AC'].width = 6
    sheet.column_dimensions['AD'].width = 6
    sheet["W4"].value = "X-Position_21"
    sheet["X4"].value = "Y-Position_21"
    sheet["Y4"].value = "X-Position_22"
    sheet["Z4"].value = "Y-Position_22"
    sheet["AA4"].value = "X-Position_23"
    sheet["AB4"].value = "Y-Position_23"
    sheet["AC4"].value = "X-Position_24"
    sheet["AD4"].value = "Y-Position_24"

    #Pos.014 TieRod
    ws.merge_cells('AE3:AF3')
    ws.merge_cells('AG3:AI3')
    #AJ3
    sheet.column_dimensions['AJ'].width = 6
    sheet["AE3"].value = "Interface"
    sheet["AG3"].value = "Shape"
    sheet["AJ3"].value = "Length"

    sheet["AE4"].value = "Standard"
    sheet["AF4"].value = "Blind (distance TR=100)"
    sheet["AG4"].value = "Conical"
    sheet["AH4"].value = " "
    sheet["AI4"].value = "Cylindrical"
    sheet["AJ4"].value = "Total-length [mm]"

    #Pos.16
    sheet.column_dimensions['AK'].width = 10
    sheet["AK3"].value = "Hub"

    sheet["AK4"].value = "Height Hub [mm]"

    #Pos.2x Stud
    ws.merge_cells('AL3:AN3')
    #AO3
    sheet["AL3"].value = "Shape"
    sheet.column_dimensions['AO'].width = 6
    sheet["AO3"].value = "Length"

    sheet["AL4"].value = "Conical"
    sheet["AM4"].value = " "
    sheet["AN4"].value = "Cylindrical"
    sheet["AO4"].value = "Length [mm]"

    #Pos.28 DDU
    #AP3-empty
    sheet["AQ3"].value = "Pos22"
    sheet["AR3"].value = "Pos20"
    #AS3-empty
    #AT3-empty
    sheet["AU3"].value = "Adapter CP"

    sheet.column_dimensions['AQ'].width = 6
    sheet.column_dimensions['AR'].width = 6
    sheet.column_dimensions['AU'].width = 7
    sheet["AP4"].value = "Keypad length"
    sheet["AQ4"].value = "InputRod length [mm]"
    sheet["AR4"].value = "Plunger length [mm]"
    sheet["AS4"].value = "Total-length [mm]"
    sheet["AT4"].value = "Enhanced Key"
    sheet["AU4"].value = "Adapter [mm]"

    #Pos.034 ValveBody
    sheet["AV3"].value = "Ø A"
    sheet["AW3"].value = " "
    sheet["AX3"].value = "Design"

    sheet.column_dimensions['AX'].width = 6
    sheet["AV4"].value = "Ø A [mm]"
    sheet["AW4"].value = " "
    sheet["AX4"].value = "Height [mm]"

    #Pos.36
    ws.merge_cells('AY3:BD3')
    sheet["AY3"].value = "Plunger Plate"

    sheet["AY4"].value = " "
    sheet["AZ4"].value = "CamProg PlungerPlate"
    sheet["BA4"].value = "Inside-Ø D [mm]"
    sheet["BB4"].value = "Outside-Ø D [mm]"
    sheet["BC4"].value = "Length L0"
    sheet["BD4"].value = "Length L1"

    #Pos.40
    #BE4-empty
    sheet.column_dimensions['BE'].width = 10
    sheet["BE4"].value = "CamProg TieRod-O-Ring"

    #Pos.42 Push Rod assy
    #BF-empty
    sheet.column_dimensions['BG'].width = 6
    sheet["BG3"].value = "Pos.60"
    #BH4-empty

    sheet["BF4"].value = "Length BallRod"
    sheet["BG4"].value = "Length_C_PushRod Penetration"
    sheet["BH4"].value = " "

    #Pos.50
    ws.merge_cells('BI3:BM3')
    sheet["BI3"].value = "Cover"

    sheet.column_dimensions['BI'].width = 10
    sheet["BI4"].value = "CoverDesign"
    sheet["BJ4"].value = "PlugDesign"
    sheet["BK4"].value = "CoverCoating"
    sheet["BL4"].value = "CoverHeight"
    sheet["BM4"].value = " "

    #Pos.058 PowerPack
    ws.merge_cells('BN3:BQ3')
    ws.merge_cells('BR3:BS3')
    ws.merge_cells('BT3:BU3')
    sheet["BN3"].value = "Design"
    sheet["BR3"].value = "ECU"
    sheet["BT3"].value = "Motor"

    sheet["BN4"].value = "700 - Base Connector"
    sheet["BO4"].value = "710 - Shrink connector"
    sheet["BP4"].value = " "
    sheet["BQ4"].value = " "
    sheet["BR4"].value = "Connector Type"
    sheet.column_dimensions['BS'].width = 6
    sheet["BS4"].value = "ECU Orientation"
    sheet["BT4"].value = "Epsilon Pos.999"
    sheet["BU4"].value = "Motor Size S1=10, S1+=11, S2=20, Brose-Motor=5"

    #Pos.82
    ws.merge_cells('BV3:BX3')
    sheet["BV3"].value = "PFS"

    sheet["BV4"].value = "CamProg PedalInterface"
    sheet["BW4"].value = "Min moment"
    sheet["BX4"].value = "Max moment"

    #Pos.083
    ws.merge_cells('BY3:CA3')
    sheet["BY3"].value = "Firewall Gasket"

    sheet["BY4"].value = "Distance ToolBore"
    sheet["BZ4"].value = "MountingPosition / ToolLifterAngle"
    sheet["CA4"].value = "NonSticky (NS) / Sticky (S) / DoubleSticky (DS)"

    #Pos.084
    ws.merge_cells('CB3:CC3')
    sheet["CB3"].value = "Spacer/InterfacePlate"

    sheet["CB4"].value = "Spacer ProgNumber"
    sheet.column_dimensions['CB'].width = 9
    sheet.column_dimensions['CC'].width = 9
    sheet["CC4"].value = "InterfacePlate ProgNumber"

    #Pos.088
    ws.merge_cells('CD3:CG3')
    sheet["CD3"].value = "Firewall Gasket"

    sheet["CD4"].value = "Distance ToolBore"
    sheet["CE4"].value = "MountingPosition / ToolLifterAngle"
    sheet["CF4"].value = "Tool Gasket assembly"
    sheet["CG4"].value = "NonSticky (NS) / Sticky (S) / DoubleSticky (DS)"

    #Pos.094
    ws.merge_cells('CH3:CL3')
    sheet["CH3"].value = "TMC"

    sheet["CH4"].value = "CamProg TMC"
    sheet["CI4"].value = "CheckDMC_ScanProg"
    sheet["CJ4"].value = " "
    sheet["CK4"].value = " "
    sheet.column_dimensions['CL'].width = 12
    sheet["CL4"].value = "TMC_DMC"

    #Pos.97
    #CM3-empty
    sheet.column_dimensions['CM'].width = 10
    sheet["CM4"].value = "Protective Cap ECU"

    #Pos.499
    ws.merge_cells('CN3:CO3')
    sheet["CN3"].value = "Label Marking"

    sheet["CN4"].value = " "
    sheet.column_dimensions['CO'].width = 20
    sheet["CO4"].value = "LabelFileName"

    #-LS - Pos.32
    ws.merge_cells('CP3:CS3')
    #CP3-empty

    sheet["CP4"].value = "1264500100"
    sheet["CQ4"].value = "1264500101"
    sheet["CR4"].value = "1264500102"
    sheet["CS4"].value = "1264500105"

    #Testing - DryTest
    ws.merge_cells('CT2:CU3')
    sheet["CT2"].value = "DryTest"
    
    sheet.column_dimensions['CU'].width = 8
    sheet["CT4"].value = "LipS-SB Plug"
    sheet["CU4"].value = "Typedata Typegroup - nr from TS"

    #zapis sformatowanego pliku do .xlsx
    wb.save(path)

def get_data_from_xml(path_02, path_07, path_165, path_53):
 
    #puste listy i zadeklarowane wcześniej wartości komórek
    Stud_pos = []
    positions = ["StudPos21X","StudPos21Y", "StudPos22X","StudPos22Y", "StudPos23X","StudPos23Y","StudPos24X","StudPos24Y"]
    for p in positions:
        Stud_pos.append('-')
    
    FirewallGasket_pos = []
    positions_FG = ["ToolBoreDistance", "Gasket_Type", "Toolplate", "ToolLifterAngle"]
    for p_FG in positions_FG:
        FirewallGasket_pos.append('-')
        
    FirewallGasket83_pos = []
    positions_FG83 = ["ToolBoreDistance", "Gasket_Type", "ToolLifterAngle"]
    for p_FG83 in positions_FG83:
        FirewallGasket83_pos.append('-')
    global Pos083_ToolBoreDistance
    global Pos083_Gasket_Type
    global Pos083_ToolLifterAngle
    Pos083_ToolBoreDistance = '-'
    Pos083_Gasket_Type = '-'
    Pos083_ToolLifterAngle = '-'

    MCRA_pos = []
    positions_MCRA = ["TMC_DMC_Content", "CheckTMC_CamProg", "CheckDMC_ScanProg"]
    for p_MCRA in positions_MCRA:
        MCRA_pos.append('-')
        
    global Pos994_LabelFileName
    Pos994_LabelFileName = "no label"
    
    global MT_TLA
    MT_TLA = '-'
    
    global MT_St65Spacer
    MT_St65Spacer = '-'
    
    global MT_St65Gasket
    MT_St65Gasket = '-'
    
    global BTR1
    global BTR2
    BTR1 = '-'
    BTR2 = '-'
    
    global Pos97_Protective_Cap_ECU
    Pos97_Protective_Cap_ECU = '-'
    
    global GH_ZAM
    GH_ZAM = '-'
    
    global Pos085_InterfaceP
    Pos085_InterfaceP = '-'
    
    global Pos084_Spacer
    Pos084_Spacer = '-'
   

    #przeszukiwanie folderu cell02
    tree_02 = ET.parse(path_02)
    root_02 = tree_02.getroot()
    
    for Type in root_02.iter('Type'):
        PN = Type.get('name')
        DS = Type.get('Desc')
        
        for Component in root_02.iter('Component'):
            CName = Component.get('name')
            if CName == 'Pos010_GearHousing':
                 for ComponentNo in Component.iter('ComponentNo'):
                    GH = ComponentNo.get('name')
                    for CompAttrVal in Component.iter('CompAttrVal'):
                        GH_AttrVal = CompAttrVal.get('name') 
                        if GH_AttrVal == 'GearHousingDesign':
                            GH_Design = CompAttrVal.get('Value')
                            if GH_Design == 'mirrored':
                                GH_Design_standard = '-'
                                GH_Design_mirrored = 'x'
                            elif GH_Design == 'standard':
                                GH_Design_standard = 'x'
                                GH_Design_mirrored = '-'

                        if GH_AttrVal == 'TieRodDistance':
                            GH_Distance = CompAttrVal.get('Value')
                            if GH_Distance == '100':
                                GH_Distance100 = 'x'
                                GH_Distance101 = '-'
                            elif GH_Distance == '101.8':
                                GH_Distance100 = '-'
                                GH_Distance101 = 'x'

                        if GH_AttrVal == 'NumberOfStuds':
                            GH_Studs = CompAttrVal.get('Value')               
                        
                        if GH_AttrVal in positions:
                            index = positions.index(GH_AttrVal)
                            Stud_pos[index] = CompAttrVal.get('Value')
                            
                        GH_Pos21X = Stud_pos[0]
                        GH_Pos21Y = Stud_pos[1]
                        GH_Pos22X = Stud_pos[2]
                        GH_Pos22Y = Stud_pos[3]
                        GH_Pos23X = Stud_pos[4]
                        GH_Pos23Y = Stud_pos[5]
                        GH_Pos24X = Stud_pos[6]
                        GH_Pos24Y = Stud_pos[7]
                        
                        if Stud_pos[0] != '-' and Stud_pos[2] != '-' and Stud_pos[4] != '-' and Stud_pos[6] != '-':
                            GH_Typgruppe_21_23 = '-'
                            GH_Typgruppe_22_24 = '-'
                            GH_Typgruppe_21_24 = 'x'
                        elif Stud_pos[0] != '-' and Stud_pos[4] != '-':
                            GH_Typgruppe_21_23 = 'x'
                            GH_Typgruppe_22_24 = '-'
                            GH_Typgruppe_21_24 = '-'
                        elif Stud_pos[2] != '-' and Stud_pos[6] != '-':
                            GH_Typgruppe_21_23 = '-'
                            GH_Typgruppe_22_24 = 'x'
                            GH_Typgruppe_21_24 = '-'

            if CName == 'Pos058_PowerPack':
                 for ComponentNo in Component.iter('ComponentNo'):
                    PP = ComponentNo.get('name')
                    
            if CName == 'Pos082_PedalInterface':
                 for ComponentNo in Component.iter('ComponentNo'):
                    PI = ComponentNo.get('name')
                                
            if CName == 'MachineTools':
                 for ComponentNo in Component.iter('ComponentNo'):
                    MT = ComponentNo.get('name')
                    if MT == 'St53_DryTesting':
                         for CompAttrVal in Component.iter('CompAttrVal'):
                            MT_St53 = CompAttrVal.get('name')
                            if MT_St53 == 'TesterECUPlugMirror':
                                MT_St53Mirror = CompAttrVal.get('Value') 
                            elif MT_St53 == 'TesterECUPlugAngle':
                                MT_St53Angle = CompAttrVal.get('Value')         
                    elif MT == 'St73_FirewallGasket':
                        for CompAttrVal in Component.iter('CompAttrVal'):
                            MT_St73 = CompAttrVal.get('name')
                            if MT_St73 == 'Mounting':
                                global MT_St73Mounting
                                MT_St73Mounting = CompAttrVal.get('Value') 
            
            if CName == 'Pos088_FirewallGasket':
                for ComponentNo in Component.iter('ComponentNo'):
                    Pos088 = ComponentNo.get('name')
                    if Pos088 == 'MountingPosition':
                        for CompAttrVal in Component.iter('CompAttrVal'):
                            TLA = CompAttrVal.get('name')
                            if TLA == 'ToolLifterAngle':
                                MT_TLA = CompAttrVal.get('Value') 
                                
            if CName == 'Pos014_TieRod':
                 for ComponentNo in Component.iter('ComponentNo'):
                    Pos014 = ComponentNo.get('name')  
                    for CompAttrVal in Component.iter('CompAttrVal'):
                        Pos014_AttrVal = CompAttrVal.get('name') 
                        if Pos014_AttrVal == 'TieRodInterface':
                            Pos014_Interface = CompAttrVal.get('Value')
                            if Pos014_Interface == 'standard':
                                Pos014_Interface_standard = 'x'
                                Pos014_Interface_blind = '-'
                            elif Pos014_Interface == 'blind':
                                Pos014_Interface_standard = '-'
                                Pos014_Interface_blind = 'x'
                        if Pos014_AttrVal == 'TieRodLength':
                            Pos014_Lenght = CompAttrVal.get('Value')
                        if Pos014_AttrVal == 'TieRodShape':
                            Pos014_Shape = CompAttrVal.get('Value')
                            if Pos014_Shape == 'conical':
                                Pos014_Shape_conical = 'x'
                                Pos014_Shape_cylindrical = '-'
                            elif Pos014_Shape == 'cylindrical':
                                Pos014_Shape_conical = '-'
                                Pos014_Shape_cylindrical = 'x'
            
            if CName == 'Pos016_GearHub':
                for ComponentNo in Component.iter('ComponentNo'):
                    Pos016 = ComponentNo.get('name')  
                    for CompAttrVal in Component.iter('CompAttrVal'):
                        Pos016_AttrVal = CompAttrVal.get('name') 
                        if Pos016_AttrVal == 'Height':
                            Pos016_Height = CompAttrVal.get('Value')
                            
            if CName == 'Pos020_Stud':
                 for ComponentNo in Component.iter('ComponentNo'):
                    Pos020 = ComponentNo.get('name')  
                    for CompAttrVal in Component.iter('CompAttrVal'):
                        Pos020_AttrVal = CompAttrVal.get('name') 
                        if Pos020_AttrVal == 'StudLength':
                            global Pos020_Lenght
                            Pos020_Lenght = CompAttrVal.get('Value')
                        if Pos020_AttrVal == 'StudShape':
                            Pos020_Shape = CompAttrVal.get('Value')
                            if Pos020_Shape == 'conical':
                                global Pos020_Shape_conical
                                global Pos020_Shape_cylindrical
                                Pos020_Shape_conical = 'x'
                                Pos020_Shape_cylindrical = '-'
                            elif Pos020_Shape == 'cylindrical':
                                Pos020_Shape_conical = '-'
                                Pos020_Shape_cylindrical = 'x'
            
            if CName == 'Pos028_DriverDemandUnit':
                 for ComponentNo in Component.iter('ComponentNo'):
                    Pos028 = ComponentNo.get('name')  
                    for CompAttrVal in Component.iter('CompAttrVal'):
                        Pos028_AttrVal = CompAttrVal.get('name') 
                        if Pos028_AttrVal == 'InPutRodLength':
                            Pos28_InputRodLength = CompAttrVal.get('Value')
                            
            if CName == 'Pos034_ValveBody':
                 for ComponentNo in Component.iter('ComponentNo'):
                    Pos034 = ComponentNo.get('name')  
                    for CompAttrVal in Component.iter('CompAttrVal'):
                        Pos034_AttrVal = CompAttrVal.get('name') 
                        if Pos034_AttrVal == 'CenterDiameter':
                            global Pos034_CenterDiameter
                            Pos034_CenterDiameter = CompAttrVal.get('Value')
                        elif Pos034_AttrVal == 'Height':
                            global Pos034_Height
                            Pos034_Height = CompAttrVal.get('Value')
                            
            if CName == 'Pos036_PlungerPlate':
                 for ComponentNo in Component.iter('ComponentNo'):
                    Pos036 = ComponentNo.get('name')  
                    for CompAttrVal in Component.iter('CompAttrVal'):
                        Pos036_AttrVal = CompAttrVal.get('name') 
                        if Pos036_AttrVal == 'CheckPlungerPlate_CamProg':
                            Pos036_CheckPlungerPlate_CamProg = CompAttrVal.get('Value')
                        elif Pos036_AttrVal == 'Inside_Diameter':
                            Pos036_Inside_Diameter = CompAttrVal.get('Value')
                        elif Pos036_AttrVal == 'Outside_Diameter':
                            Pos036_Outside_Diameter = CompAttrVal.get('Value')
                        elif Pos036_AttrVal == 'Length_L0':
                            Pos036_Length_L0 = CompAttrVal.get('Value')
                        elif Pos036_AttrVal == 'Length_L1':
                            Pos036_Length_L1 = CompAttrVal.get('Value')
                            
            if CName == 'Pos040_TieRodORing':
                 for ComponentNo in Component.iter('ComponentNo'):
                    Pos040 = ComponentNo.get('name')  
                    for CompAttrVal in Component.iter('CompAttrVal'):
                        Pos040_AttrVal = CompAttrVal.get('name') 
                        if Pos040_AttrVal == 'CheckTieRodORing_CamProg':
                            Pos040_CheckTieRodORing_CamProg = CompAttrVal.get('Value')
                            
            if CName == 'Pos042_BallRod':
                 for ComponentNo in Component.iter('ComponentNo'):
                    Pos042 = ComponentNo.get('name')  
                    for CompAttrVal in Component.iter('CompAttrVal'):
                        Pos042_AttrVal = CompAttrVal.get('name') 
                        if Pos042_AttrVal == 'BallRodLength':
                            global Pos042_BallRodLength
                            Pos042_BallRodLength = CompAttrVal.get('Value')    
                        elif Pos042_AttrVal == 'C_Penetration':
                            global Pos042_C_Penetration
                            Pos042_C_Penetration = CompAttrVal.get('Value')  
                            
            if CName == 'Pos050_Cover':
                 for ComponentNo in Component.iter('ComponentNo'):
                    Pos050 = ComponentNo.get('name')  
                    for CompAttrVal in Component.iter('CompAttrVal'):
                        Pos050_AttrVal = CompAttrVal.get('name') 
                        if Pos050_AttrVal == 'PlugDesign':
                            global Pos050_PlugDesign
                            Pos050_PlugDesign = CompAttrVal.get('Value')   
            
            if CName == 'Pos058_PowerPack':
                 for ComponentNo in Component.iter('ComponentNo'):
                    Pos058 = ComponentNo.get('name')  
                    for CompAttrVal in Component.iter('CompAttrVal'):
                        Pos058_AttrVal = CompAttrVal.get('name') 
                        if Pos058_AttrVal == 'ConnectorType':
                            Pos058_ConnectorType = CompAttrVal.get('Value') 
                            if Pos058_ConnectorType == '700':
                                Pos058_700BaseConnector = 'x'
                                Pos058_710ShrinkConnector = '-'
                            elif Pos058_ConnectorType == '710':
                                Pos058_700BaseConnector = '-'
                                Pos058_710ShrinkConnector = 'x'
                        elif Pos058_AttrVal == 'MotorSize':
                            Pos058_MotorSize = CompAttrVal.get('Value')  
                        elif Pos058_AttrVal == 'ECUOrientation':
                            Pos058_ECUOrientation = CompAttrVal.get('Value')
                        elif Pos058_AttrVal == 'MotorOrientation_epsilon':
                            Pos058_MotorOrientation_epsilon = CompAttrVal.get('Value') 
                            
            if CName == 'Pos088_FirewallGasket':
                 for ComponentNo in Component.iter('ComponentNo'):
                    Pos088 = ComponentNo.get('name')  
                    for CompAttrVal in Component.iter('CompAttrVal'):
                        Pos088_AttrVal = CompAttrVal.get('name') 
                        
                        if Pos088_AttrVal in positions_FG:
                            index = positions_FG.index(Pos088_AttrVal)
                            FirewallGasket_pos[index] = CompAttrVal.get('Value')
                        
                        global Pos088_ToolBoreDistance
                        global Pos088_Gasket_Type
                        global Pos088_Toolplate
                        global Pos088_ToolLifterAngle
                        
                        Pos088_ToolBoreDistance = FirewallGasket_pos[0]
                        Pos088_Gasket_Type = FirewallGasket_pos[1]
                        Pos088_Toolplate = FirewallGasket_pos[2]
                        Pos088_ToolLifterAngle = FirewallGasket_pos[3]
            
            if CName == 'Pos094_MasterCylReservoirAssy':
                 for ComponentNo in Component.iter('ComponentNo'):
                    Pos094 = ComponentNo.get('name')  
                    for CompAttrVal in Component.iter('CompAttrVal'):
                        Pos094_AttrVal = CompAttrVal.get('name') 
                        
                        if Pos094_AttrVal in positions_MCRA:
                            index = positions_MCRA.index(Pos094_AttrVal)
                            MCRA_pos[index] = CompAttrVal.get('Value')
                        
                        global Pos094_TMC_DMC_Content
                        global Pos094_CheckTMC_CamProg
                        global Pos094_CheckDMC_ScanProg
                        
                        Pos094_TMC_DMC_Content = MCRA_pos[0]
                        Pos094_CheckTMC_CamProg = MCRA_pos[1]
                        Pos094_CheckDMC_ScanProg = MCRA_pos[2]
                        
            if CName == 'Pos032_LinearPositionSensor':
                 for ComponentNo in Component.iter('ComponentNo'):
                    Pos032 = ComponentNo.get('name')  
                    global Pos032_100
                    global Pos032_101
                    global Pos032_102
                    global Pos032_105
                    global Testing_LipS_SB_Plug
                    if Pos032 == '1264500100':
                        Pos032_100 = 'x'
                        Pos032_101 = '-'
                        Pos032_102 = '-'
                        Pos032_105 = '-'
                        Testing_LipS_SB_Plug = '01'
                    elif Pos032 == '1264500101':
                        Pos032_100 = '-'
                        Pos032_101 = 'x'
                        Pos032_102 = '-'
                        Pos032_105 = '-'
                        Testing_LipS_SB_Plug = '02'
                    elif Pos032 == '1264500102':
                        Pos032_100 = '-'
                        Pos032_101 = '-'
                        Pos032_102 = 'x'
                        Pos032_105 = '-'
                        Testing_LipS_SB_Plug = '03'
                    elif Pos032 == '1264500105':
                        Pos032_100 = '-'
                        Pos032_101 = '-'
                        Pos032_102 = '-'
                        Pos032_105 = 'x'
                        Testing_LipS_SB_Plug = '01'
                        
            if CName == 'Pos994_L_LabelFileName':
                 for ComponentNo in Component.iter('ComponentNo'):
                    Pos994 = ComponentNo.get('name')  
                    for CompAttrVal in Component.iter('CompAttrVal'):
                        Pos994_AttrVal = CompAttrVal.get('name') 
                        if Pos994_AttrVal == 'LabelFileName':
                            Pos994_LabelFileName = CompAttrVal.get('Value') 
                            
            if CName == 'Pos083_FirewallGasket':
                 for ComponentNo in Component.iter('ComponentNo'):
                    Pos083 = ComponentNo.get('name')  
                    for CompAttrVal in Component.iter('CompAttrVal'):
                        Pos083_AttrVal = CompAttrVal.get('name') 
                        if Pos083_AttrVal in positions_FG83:
                            index = positions_FG83.index(Pos083_AttrVal)
                            FirewallGasket83_pos[index] = CompAttrVal.get('Value')
                        Pos083_ToolBoreDistance = FirewallGasket83_pos[0]
                        Pos083_Gasket_Type = FirewallGasket83_pos[1]
                        Pos083_ToolLifterAngle = FirewallGasket83_pos[2]
                        
            if CName == 'Pos097_ProtectivCap':
                Pos97_Protective_Cap_ECU = 'x'
        
    
        #przeszukiwanie folderu cell07
        tree_07 = ET.parse(path_07)
        root_07 = tree_07.getroot()

        for Type in root_07.iter('Type'):
            PN_07 = Type.get('name')
            if PN_07 == PN:
                for Cell in root_07.iter('Cell'):
                    for MeasurementPoint in root_07.iter('MeasurementPoint'):
                        MPoint = MeasurementPoint.get('name')
                        if MPoint == 'BTR_RivetingJob1':
                            for Feature in MeasurementPoint.iter('Feature'):
                                BTR1test = Feature.get('Value')
                                BTR1 = BTR1test.split('E')[0]
                        elif MPoint == 'BTR_RivetingJob2':
                            for Feature in MeasurementPoint.iter('Feature'):
                                BTR2test = Feature.get('Value')
                                BTR2 = BTR2test.split('E')[0]
        
                                        
        #przeszukiwanie folderu cell165
        tree_165 = ET.parse(path_165)
        root_165 = tree_165.getroot()

        for Type in root_165.iter('Type'):
            PN_165 = Type.get('name')
            if PN_165 == PN:
                for Component in root_165.iter('Component'):
                    CName = Component.get('name')
                    if CName == 'Pos082_PedalInterface':
                        for ComponentNo in Component.iter('ComponentNo'):
                            Pos082 = ComponentNo.get('name')  
                            for CompAttrVal in Component.iter('CompAttrVal'):
                                Pos082_AttrVal = CompAttrVal.get('name') 
                                if Pos082_AttrVal == 'CheckPedalInterface_CamProg':
                                    Pos082_CheckPedalInterface_CamProg = CompAttrVal.get('Value') 
                                elif Pos082_AttrVal == 'Pedalinterface_Min_Moment':
                                    Pos082_Min_Moment = CompAttrVal.get('Value')
                                elif Pos082_AttrVal == 'Pedalinterface_Max_Moment':
                                    Pos082_Max_Moment = CompAttrVal.get('Value')
                    
                    if CName == 'MachineTools':
                        for ComponentNo in Component.iter('ComponentNo'):
                            St65 = ComponentNo.get('name')
                            if St65 == 'St65_CameraCheck':
                                for CompAttrVal in Component.iter('CompAttrVal'):
                                    St65_AttrVal = CompAttrVal.get('name') 
                                    if St65_AttrVal == 'CamProg_Gasket':
                                        MT_St65Gasket = CompAttrVal.get('Value')
                                    elif St65_AttrVal == 'CamProg_Spacer':
                                        MT_St65Spacer = CompAttrVal.get('Value')
                    
                    if CName == 'Pos084_Spacer':
                        for ComponentNo in Component.iter('ComponentNo'):
                            Pos084 = ComponentNo.get('name')
                            for CompAttrVal in Component.iter('CompAttrVal'):
                                Pos084_Spacer = CompAttrVal.get('Value')
                                
                    if CName == 'Pos085_InterfacePlate':
                        for ComponentNo in Component.iter('ComponentNo'):
                            Pos085 = ComponentNo.get('name')
                            for CompAttrVal in Component.iter('CompAttrVal'):
                                Pos085_InterfaceP = CompAttrVal.get('Value')
                                
                                
                    if CName == 'Pos050_Cover':
                        for ComponentNo in Component.iter('ComponentNo'):
                            Pos050 = ComponentNo.get('name')
                            for CompAttrVal in Component.iter('CompAttrVal'):
                                Pos050_AttrVal = CompAttrVal.get('name')
                                if Pos050_AttrVal == 'CoverDesign':
                                    global Pos050_CoverDesign
                                    Pos050_CoverDesign = CompAttrVal.get('Value')
                                elif Pos050_AttrVal == 'PlugDesign':
                                    Pos050_PlugDesign = CompAttrVal.get('Value')
                                elif Pos050_AttrVal == 'CoverCoating':
                                    global Pos050_CoverCoating
                                    Pos050_CoverCoating = CompAttrVal.get('Value')
                                    if  Pos050_CoverCoating == 'ZAM':
                                        GH_ZAM = 'x'
                                elif Pos050_AttrVal == 'CoverHeight':
                                    global Pos050_CoverHeight
                                    Pos050_CoverHeight = CompAttrVal.get('Value')
                  
        #przeszukiwanie folderu cell53
        tree_53 = ET.parse(path_53)
        root_53 = tree_53.getroot()

        for Type in root_53.iter('Type'):
            PN_53 = Type.get('name')
            if PN_53 == PN:
                for Cell in root_53.iter('Cell'):
                    for Module in root_53.iter('Module'):
                        for Step in root_53.iter('Step'):
                            for MeasurementPoint in root_53.iter('MeasurementPoint'):
                                MPoint = MeasurementPoint.get('name')
                                if MPoint == 'Testtypegroup_GLDf':
                                    for Feature in MeasurementPoint.iter('Feature'):
                                        Testing_Typegroup_nr = Feature.get('String')[-4:]
                                      
    return PN, GH, GH_Design_standard, GH_Design_mirrored, GH_Distance100, GH_Distance101, GH_Studs, GH_Typgruppe_21_23, GH_Typgruppe_22_24, GH_Typgruppe_21_24, GH_Pos21X, GH_Pos21Y, GH_Pos22X, GH_Pos22Y, GH_Pos23X, GH_Pos23Y, GH_Pos24X, GH_Pos24Y, PP, DS, PI, MT_St53Mirror, MT_St53Angle, MT_St73Mounting, MT_TLA, Pos014_Interface_standard, Pos014_Interface_blind, Pos014_Shape_conical, Pos014_Shape_cylindrical, Pos014_Lenght, Pos016_Height, Pos020_Lenght, Pos020_Shape_conical, Pos020_Shape_cylindrical, Pos28_InputRodLength, Pos034_CenterDiameter, Pos034_Height, Pos036_CheckPlungerPlate_CamProg, Pos036_Inside_Diameter, Pos036_Outside_Diameter, Pos036_Length_L0, Pos036_Length_L1, Pos040_CheckTieRodORing_CamProg, Pos042_BallRodLength, Pos042_C_Penetration, Pos058_ConnectorType, Pos058_700BaseConnector, Pos058_710ShrinkConnector, Pos058_MotorSize, Pos058_ECUOrientation, Pos058_MotorOrientation_epsilon, Pos082_CheckPedalInterface_CamProg, Pos082_Min_Moment, Pos082_Max_Moment, Pos088_ToolBoreDistance, Pos088_Gasket_Type, Pos088_Toolplate, Pos088_ToolLifterAngle, Pos094_TMC_DMC_Content, Pos094_CheckTMC_CamProg, Pos094_CheckDMC_ScanProg, Pos032_100, Pos032_101, Pos032_102, Pos032_105, Pos994_LabelFileName, MT_St65Spacer, MT_St65Gasket, BTR1, BTR2, Pos083_ToolBoreDistance, Pos083_Gasket_Type, Pos083_ToolLifterAngle, Pos97_Protective_Cap_ECU, GH_ZAM, Testing_LipS_SB_Plug, Pos084_Spacer, Pos085_InterfaceP, Pos050_CoverDesign, Pos050_PlugDesign, Pos050_CoverCoating, Pos050_CoverHeight, Testing_Typegroup_nr                  

PN_list = []
GH_list = []
GH_Design_standard_list = []
GH_Design_mirrored_list = []
GH_ZAM_list = []
GH_Distance100_list = []
GH_Distance101_list = []
GH_RivProg_BTR1_list = []
GH_RivProg_BTR2_list = []
GH_Studs_list = []
GH_Typgruppe_21_23_list = []
GH_Typgruppe_22_24_list = []
GH_Typgruppe_21_24_list = []
GH_Pos21X_list = []
GH_Pos21Y_list = []
GH_Pos22X_list = []
GH_Pos22Y_list = []
GH_Pos23X_list = []
GH_Pos23Y_list = []
GH_Pos24X_list = []
GH_Pos24Y_list = []
PP_list = []
DS_list = []
PI_list = []
MT_St53Mirror_list = []
MT_St53Angle_list = []
MT_St65Gasket_list = []
MT_St65Spacer_list = []
MT_St73Mounting_list = []
MT_TLA_list = []
Pos014_Interface_standard_list = []
Pos014_Interface_blind_list = []
Pos014_Shape_conical_list = []
Pos014_Shape_MAThread_list = []
Pos014_Shape_cylindrical_list = []
Pos014_Lenght_list = []
Pos016_Height_list = []
Pos020_Lenght_list = []
Pos020_Shape_conical_list = []
Pos020_Shape_MAThread_list = []
Pos020_Shape_cylindrical_list = []
Pos28_KeypadLength_list = []
Pos28_InputRodLength_list = []
Pos28_PlungerLength_list = []
Pos28_TotalLength_list = []
Pos28_EnhancedKey_list = []
Pos28_AdapterCP_list = []
Pos034_CenterDiameter_list = []
Pos034_Height_list = []
Pos036_PlungerPlateTyp_list = []
Pos036_CheckPlungerPlate_CamProg_list = []
Pos036_Inside_Diameter_list = []
Pos036_Outside_Diameter_list = []
Pos036_Length_L0_list = []
Pos036_Length_L1_list = []
Pos040_CheckTieRodORing_CamProg_list = []
Pos042_BallRodLength_list = []
Pos042_C_Penetration_list = []
Pos042_OffsetPushRodPenetration_list = []
Pos050_CoverDesign_list = []
Pos050_PlugDesign_list = []
Pos050_CoverCoating_list = []
Pos050_CoverHeight_list = []
Pos058_700BaseConnector_list = []
Pos058_710ShrinkConnector_list = []
Pos058_StandardPP_list = []
Pos058_EVOPP_list = []
Pos058_ConnectorType_list = []
Pos058_MotorSize_list = []
Pos058_ECUOrientation_list = []
Pos058_MotorOrientation_epsilon_list = []
Pos082_CheckPedalInterface_CamProg_list = []
Pos082_Min_Moment_list = []
Pos082_Max_Moment_list = []
Pos083_DistanceToolBore_list = []
Pos083_MountingPosition_list = []
Pos083_Sticky_list = []
Pos084_Spacer_list  = []
Pos085_InterfacePlate_list = []
Pos088_ToolBoreDistance_list = []
Pos088_Gasket_Type_list = []
Pos088_Toolplate_list = []
Pos088_ToolLifterAngle_list = []
Pos094_CheckTMC_CamProg_list = []
Pos094_CheckDMC_ScanProg_list = []
Pos094_Protection_Cover_Label_list = []
Pos094_TMC_DMC_Content_list = []
Pos97_Protective_Cap_ECU_list = []
Pos499_Res2Abrasion_list = []
Pos994_LabelFileName_list = []
Pos032_100_list = []
Pos032_101_list = []
Pos032_102_list = []
Pos032_105_list = []
Testing_LipS_SB_Plug_list = []
Testing_Typegroup_nr_list = []
Empty_list = []


#generowanie DF - określanie ścieżek do folderów
path_02 = './TD/cell02/td/cell'
path_07 = './TD/cell07/td/cell'
path_165 = './TD/cell165/td/cell'
path_53 = './TD/cell53/td/cell'

#zabezpieczenie - jakich plików nie brać pod uwagę
excl_list = ['GHOST', 'serialno', 'supplement', 'Backup', 'MASTER', '.', 'CHECK', 'bak']
for filename_02 in os.listdir(path_02):
    PartNumber = filename_02.split('/')[0].split('_')[0]
    for filename_07 in os.listdir(path_07):
        if filename_07.startswith(PartNumber):
            for filename_165 in os.listdir(path_165):
                if filename_165.startswith(PartNumber):
                    for filename_53 in os.listdir(path_53):
                        if filename_53.startswith(PartNumber):

                            if all([element not in filename_02 and filename_07 and filename_165 and filename_53 for element in excl_list]):
                                fullname_02 = os.path.join(path_02, filename_02)
                                fullname_07 = os.path.join(path_07, filename_07)
                                fullname_165 = os.path.join(path_165, filename_165)
                                fullname_53 = os.path.join(path_53, filename_53)

                                #wywoływanie funkcji get_data_from_xml
                                PN, GH, GH_Design_standard, GH_Design_mirrored, GH_Distance100, GH_Distance101, GH_Studs, GH_Typgruppe_21_23, GH_Typgruppe_22_24, GH_Typgruppe_21_24, GH_Pos21X, GH_Pos21Y, GH_Pos22X, GH_Pos22Y, GH_Pos23X, GH_Pos23Y, GH_Pos24X, GH_Pos24Y, PP, DS, PI, MT_St53Mirror, MT_St53Angle, MT_St73Mounting, MT_TLA, Pos014_Interface_standard, Pos014_Interface_blind, Pos014_Shape_conical, Pos014_Shape_cylindrical, Pos014_Lenght, Pos016_Height, Pos020_Lenght, Pos020_Shape_conical, Pos020_Shape_cylindrical, Pos28_InputRodLength, Pos034_CenterDiameter, Pos034_Height, Pos036_CheckPlungerPlate_CamProg, Pos036_Inside_Diameter, Pos036_Outside_Diameter, Pos036_Length_L0, Pos036_Length_L1, Pos040_CheckTieRodORing_CamProg, Pos042_BallRodLength, Pos042_C_Penetration, Pos058_ConnectorType, Pos058_700BaseConnector, Pos058_710ShrinkConnector, Pos058_MotorSize, Pos058_ECUOrientation, Pos058_MotorOrientation_epsilon, Pos082_CheckPedalInterface_CamProg, Pos082_Min_Moment, Pos082_Max_Moment, Pos088_ToolBoreDistance, Pos088_Gasket_Type, Pos088_Toolplate, Pos088_ToolLifterAngle, Pos094_TMC_DMC_Content, Pos094_CheckTMC_CamProg, Pos094_CheckDMC_ScanProg, Pos032_100, Pos032_101, Pos032_102, Pos032_105, Pos994_LabelFileName, MT_St65Spacer, MT_St65Gasket, BTR1, BTR2, Pos083_ToolBoreDistance, Pos083_Gasket_Type, Pos083_ToolLifterAngle, Pos97_Protective_Cap_ECU, GH_ZAM, Testing_LipS_SB_Plug, Pos084_Spacer, Pos085_InterfaceP, Pos050_CoverDesign, Pos050_PlugDesign, Pos050_CoverCoating, Pos050_CoverHeight, Testing_Typegroup_nr = get_data_from_xml(fullname_02, fullname_07, fullname_165, fullname_53)

                                #dodawanie wartości do list
                                PN_list.append(PN)
                                GH_list.append(GH)
                                GH_Design_standard_list.append(GH_Design_standard)
                                GH_Design_mirrored_list.append(GH_Design_mirrored)
                                GH_Distance100_list.append(GH_Distance100)
                                GH_Distance101_list.append(GH_Distance101)
                                GH_Studs_list.append(GH_Studs)
                                GH_Typgruppe_21_23_list.append(GH_Typgruppe_21_23)
                                GH_Typgruppe_22_24_list.append(GH_Typgruppe_22_24)
                                GH_Typgruppe_21_24_list.append(GH_Typgruppe_21_24)
                                GH_Pos21X_list.append(GH_Pos21X)
                                GH_Pos21Y_list.append(GH_Pos21Y)
                                GH_Pos22X_list.append(GH_Pos22X)
                                GH_Pos22Y_list.append(GH_Pos22Y)
                                GH_Pos23X_list.append(GH_Pos23X)
                                GH_Pos23Y_list.append(GH_Pos23Y)
                                GH_Pos24X_list.append(GH_Pos24X)
                                GH_Pos24Y_list.append(GH_Pos24Y)
                                PP_list.append(PP)
                                DS_list.append(DS)
                                PI_list.append(PI)
                                MT_St53Mirror_list.append(MT_St53Mirror)
                                MT_St53Angle_list.append(MT_St53Angle)
                                MT_St73Mounting_list.append(MT_St73Mounting)
                                MT_TLA_list.append(MT_TLA)
                                Pos014_Interface_standard_list.append(Pos014_Interface_standard)
                                Pos014_Interface_blind_list.append(Pos014_Interface_blind)
                                Pos014_Shape_conical_list.append(Pos014_Shape_conical)
                                Pos014_Shape_cylindrical_list.append(Pos014_Shape_cylindrical)
                                Pos014_Lenght_list.append(Pos014_Lenght)
                                Pos016_Height_list.append(Pos016_Height)
                                Pos020_Lenght_list.append(Pos020_Lenght)
                                Pos020_Shape_conical_list.append(Pos020_Shape_conical)
                                Pos020_Shape_cylindrical_list.append(Pos020_Shape_cylindrical)
                                Pos28_InputRodLength_list.append(Pos28_InputRodLength)
                                Pos034_CenterDiameter_list.append(Pos034_CenterDiameter)
                                Pos034_Height_list.append(Pos034_Height)
                                Pos036_CheckPlungerPlate_CamProg_list.append(Pos036_CheckPlungerPlate_CamProg)
                                Pos036_Inside_Diameter_list.append(Pos036_Inside_Diameter)
                                Pos036_Outside_Diameter_list.append(Pos036_Outside_Diameter)
                                Pos036_Length_L0_list.append(Pos036_Length_L0)
                                Pos036_Length_L1_list.append(Pos036_Length_L1)
                                Pos040_CheckTieRodORing_CamProg_list.append(Pos040_CheckTieRodORing_CamProg)
                                Pos042_BallRodLength_list.append(Pos042_BallRodLength)
                                Pos042_C_Penetration_list.append(Pos042_C_Penetration)
                                Pos058_700BaseConnector_list.append(Pos058_700BaseConnector)
                                Pos058_710ShrinkConnector_list.append(Pos058_710ShrinkConnector)
                                Pos058_ConnectorType_list.append(Pos058_ConnectorType)
                                Pos058_MotorSize_list.append(Pos058_MotorSize)
                                Pos058_ECUOrientation_list.append(Pos058_ECUOrientation)
                                Pos058_MotorOrientation_epsilon_list.append(Pos058_MotorOrientation_epsilon)
                                Pos082_CheckPedalInterface_CamProg_list.append(Pos082_CheckPedalInterface_CamProg)
                                Pos082_Min_Moment_list.append(Pos082_Min_Moment)
                                Pos082_Max_Moment_list.append(Pos082_Max_Moment)
                                Pos088_ToolBoreDistance_list.append(Pos088_ToolBoreDistance)
                                Pos088_Gasket_Type_list.append(Pos088_Gasket_Type)
                                Pos088_Toolplate_list.append(Pos088_Toolplate)
                                Pos088_ToolLifterAngle_list.append(Pos088_ToolLifterAngle)
                                Pos094_CheckTMC_CamProg_list.append(Pos094_CheckTMC_CamProg)
                                Pos094_CheckDMC_ScanProg_list.append(Pos094_CheckDMC_ScanProg) 
                                Pos094_TMC_DMC_Content_list.append(Pos094_TMC_DMC_Content)
                                Pos994_LabelFileName_list.append(Pos994_LabelFileName)
                                Pos032_100_list.append(Pos032_100)
                                Pos032_101_list.append(Pos032_101)
                                Pos032_102_list.append(Pos032_102)
                                Pos032_105_list.append(Pos032_105)
                                MT_St65Spacer_list.append(MT_St65Spacer)
                                MT_St65Gasket_list.append(MT_St65Gasket)
                                GH_RivProg_BTR1_list.append(BTR1)
                                GH_RivProg_BTR2_list.append(BTR2)
                                Pos083_DistanceToolBore_list.append(Pos083_ToolBoreDistance)
                                Pos083_MountingPosition_list.append(Pos083_ToolLifterAngle)
                                Pos083_Sticky_list.append(Pos083_Gasket_Type)
                                Pos97_Protective_Cap_ECU_list.append(Pos97_Protective_Cap_ECU)
                                GH_ZAM_list.append(GH_ZAM)
                                Testing_LipS_SB_Plug_list.append(Testing_LipS_SB_Plug)
                                Pos084_Spacer_list.append(Pos084_Spacer)
                                Pos085_InterfacePlate_list.append(Pos085_InterfaceP)
                                Pos050_CoverDesign_list.append(Pos050_CoverDesign)
                                Pos050_PlugDesign_list.append(Pos050_PlugDesign)
                                Pos050_CoverCoating_list.append(Pos050_CoverCoating)
                                Pos050_CoverHeight_list.append(Pos050_CoverHeight)
                                Testing_Typegroup_nr_list.append(Testing_Typegroup_nr)


a = {
#przy uzupełnianiu trzeba nadawać inne nazwy
    'Description': DS_list,
    'Part number': PN_list,
    'Gear Housing': GH_list,
    'Power Pack': PP_list,
    'Pedal Interface': PI_list,
    'ECU PlugMirror': MT_St53Mirror_list,
    'ECU PlugAngle': MT_St53Angle_list,
    'CamProg_Gasket': MT_St65Gasket_list,
    'CamProg_Spacer': MT_St65Spacer_list,
    'Mounting ?': MT_St73Mounting_list,
    'MountingPosition/ToolLifterAngle': MT_TLA_list,
    'Standard': GH_Design_standard_list,
    'Mirrored': GH_Design_mirrored_list,
    'ZAM': GH_ZAM_list,
    '100 Bore Distance': GH_Distance100_list,
    '101,8 Bore Distance': GH_Distance101_list,
    'Program BTR 1': GH_RivProg_BTR1_list,
    'Program BTR 2': GH_RivProg_BTR2_list,
    'Number of Stud Bores': GH_Studs_list,
    '2_Studs_RobPos_21_23': GH_Typgruppe_21_23_list,
    '2_Studs_RobPos_22_24': GH_Typgruppe_22_24_list,
    '4_Studs_RobPos_21-24': GH_Typgruppe_21_24_list,
    'X-Position_21': GH_Pos21X_list,
    'Y-Position_21': GH_Pos21Y_list,
    'X-Position_22': GH_Pos22X_list,
    'Y-Position_22': GH_Pos22Y_list,
    'X-Position_23': GH_Pos23X_list,
    'Y-Position_23': GH_Pos23Y_list,
    'X-Position_24': GH_Pos24X_list,
    'Y-Position_24': GH_Pos24Y_list,
    'Standard014': Pos014_Interface_standard_list,
    'Blind (distance TR=100)': Pos014_Interface_blind_list,
    'Conical': Pos014_Shape_conical_list,
    '1': Empty_list,
    'Cylindrical': Pos014_Shape_cylindrical_list,
    'Total-length [mm]': Pos014_Lenght_list,
    'Height Hub [mm]': Pos016_Height_list,
    'Conical020': Pos020_Shape_conical_list,
    '2': Empty_list,
    'Cylindrical020': Pos020_Shape_cylindrical_list,
    'Length [mm]': Pos020_Lenght_list,
    'Keypad length': Pos28_KeypadLength_list,
    'InputRod length [mm]': Pos28_InputRodLength_list,
    'Plunger length [mm]': Pos28_PlungerLength_list,
    'Total-length28 [mm]': Pos28_TotalLength_list,
    'Enhanced Key': Pos28_EnhancedKey_list,
    'Adapter [mm]': Pos28_AdapterCP_list,
    'Ø A [mm]': Pos034_CenterDiameter_list,
    '3': Empty_list,
    'Height [mm]': Pos034_Height_list,
    '4': Empty_list,
    'CamProg PlungerPlate': Pos036_CheckPlungerPlate_CamProg_list,
    'Inside-Ø D [mm]': Pos036_Inside_Diameter_list,
    'Outside-Ø D [mm]': Pos036_Outside_Diameter_list,
    'Length L0': Pos036_Length_L0_list,
    'Length L1': Pos036_Length_L1_list,
    'CamProg TieRod-O-Ring': Pos040_CheckTieRodORing_CamProg_list,
    'Length BallRod': Pos042_BallRodLength_list,
    'Length_C_PushRod Penetration': Pos042_C_Penetration_list,
    '5': Empty_list,
    'Pos050_CoverDesign': Pos050_CoverDesign_list,
    'Pos050_PlugDesign': Pos050_PlugDesign_list,
    'Pos050_CoverCoating': Pos050_CoverCoating_list,
    'Pos050_CoverHeight': Pos050_CoverHeight_list,
    '6': Empty_list,
    '700 - Base Connector': Pos058_700BaseConnector_list,
    '710 - Shrink connector': Pos058_710ShrinkConnector_list,
    '7': Empty_list,
    '8': Empty_list,
    'Connector Type': Pos058_ConnectorType_list,
    'ECU Orientation': Pos058_ECUOrientation_list,
    'Epsilon Pos.999': Pos058_MotorOrientation_epsilon_list,
    'Motor Size S1=10, S1+=11, S2=20, Brose-Motor=5': Pos058_MotorSize_list,
    'CamProg PedalInterface': Pos082_CheckPedalInterface_CamProg_list,
    'Min moment': Pos082_Min_Moment_list,
    'Max moment': Pos082_Max_Moment_list,
    'Distance ToolBore 083': Pos083_DistanceToolBore_list,
    'MountingPosition / ToolLifterAngle 083': Pos083_MountingPosition_list,
    'NonSticky (NS) / Sticky (S) / DoubleSticky (DS) 083': Pos083_Sticky_list,
    'Spacer ProgNumber': Pos084_Spacer_list,
    'InterfacePlate ProgNumber': Pos085_InterfacePlate_list,
    'Distance ToolBore 088': Pos088_ToolBoreDistance_list,
    'MountingPosition / ToolLifterAngle 088': Pos088_ToolLifterAngle_list,
    'Tool Gasket assembly 088': Pos088_Toolplate_list,
    'NonSticky (NS) / Sticky (S) / DoubleSticky (DS) 088': Pos088_Gasket_Type_list,
    'CamProg TMC': Pos094_CheckTMC_CamProg_list, 
    'CheckDMC_ScanProg': Pos094_CheckDMC_ScanProg_list, 
    '9': Empty_list,
    '10': Empty_list,
    'TMC_DMC': Pos094_TMC_DMC_Content_list,
    'Protective Cap ECU': Pos97_Protective_Cap_ECU_list,
    '11': Empty_list,
    'LabelFileName': Pos994_LabelFileName_list,
    '1264500100': Pos032_100_list,
    '1264500101': Pos032_101_list,
    '1264500102': Pos032_102_list,
    '1264500105': Pos032_105_list,
    'LipS-SB Plug': Testing_LipS_SB_Plug_list,
    'Typedata Typegroup - nr from TS': Testing_Typegroup_nr_list,
    }

#generowanie DF
df = pd.DataFrame.from_dict(a, orient='index')
df = df.transpose()

#zapis DF do pliku
df = df.set_index('Description', drop = True)
df.to_excel("./data_frames_from_xml.xlsx", startrow=3, freeze_panes=(5,5))

#funkcja formatująca wygląd excela
formatting("./data_frames_from_xml.xlsx")
